# analyze_cases.py — ICAIL 2026
# Readability & Citation Analysis for Apex Court Judgments
#
# Accompanies: "Empirical Analysis of Judicial Writing Across Five Apex Courts"
# For: ICAIL 2026 (International Conference on Artificial Intelligence and Law)
#
# Analyzes cleaned legal text files (.txt) and generates Excel/CSV with:
# - Case metadata (title, citation, date, year, court, country)
# - Readability metrics (FK Grade, FK Ease, SMOG, avg sentence length)
#   Uses legal-aware NLTK Punkt tokenizer (~70 abbreviations: v., No., para., s., etc.)
#   and regex-based citation stripping before metrics calculation.
# - Citation counts by jurisdiction (SG, UK, AU, USA, CAN, IND, NZ, EU, OTHER)
# - Academic references count (journals, books, named treatises)
#
# Usage:
#   python analyze_cases.py --court SGCA
#   python analyze_cases.py --input path/to/cleaned_texts --output results.xlsx
#   python analyze_cases.py --input path/to/cleaned_texts --court UKSC --csv

"""
Readability & Citation Analysis for Apex Court Judgments (ICAIL 2026)

Analyzes cleaned legal documents and generates Excel/CSV with:
- Case metadata (title, citation, date, year, court, country)
- Readability metrics (FK Grade, FK Ease, SMOG, avg sentence length)
  Uses legal-aware NLTK Punkt tokenizer (~70 abbreviations: v., No., para., s., etc.)
  and regex-based citation stripping before metrics calculation.
- Citation counts by jurisdiction (SG, UK, AU, USA, CAN, IND, NZ, EU, OTHER)
- Academic references count (journals, books)

Usage:
    python analyze_cases.py --court SGCA
    python analyze_cases.py --court HCA
    python analyze_cases.py --input [folder] --output [excel_path]
"""

import os
import re
import sys
import math
import argparse
from typing import Dict, List, Tuple, Optional

import nltk
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# ============================================================================
# CONFIGURATION
# ============================================================================

# Default output directory — change to your project's output folder
OUTPUT_BASE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "output")
FINAL_BASE = os.path.join(OUTPUT_BASE, "Final Clean Text Judgments")
FINAL_BASE_2 = os.path.join(OUTPUT_BASE, "Final Clean Text Judgments 2.0")

COURT_CONFIG = {
    'SGCA': {
        'name': 'Singapore Court of Appeal',
        'input_folders': [os.path.join(FINAL_BASE, "SGCA_2005-2024 Final 1.0")],
        'country': 'SG',
    },
    'SGHC': {
        'name': 'Singapore High Court',
        'input_folders': [os.path.join(FINAL_BASE, "SGHC_2005-2024 Final 1.0")],
        'country': 'SG',
    },
    'UKHL': {
        'name': 'UK House of Lords',
        'input_folders': [os.path.join(FINAL_BASE, "HL_1996-2009 Final 1.0")],
        'country': 'UK',
    },
    'UKSC': {
        'name': 'UK Supreme Court',
        'input_folders': [os.path.join(FINAL_BASE, "UKSC_2009-2024 Final 1.0")],
        'country': 'UK',
    },
    'HCA': {
        'name': 'High Court of Australia',
        'input_folders': [os.path.join(FINAL_BASE_2, "HCA_2000-2025")],
        'country': 'AU',
    },
    'SGDC': {
        'name': 'Singapore District Court',
        'input_folders': [os.path.join(FINAL_BASE, "SGDC_2005-2024 Final 1.0")],
        'country': 'SG',
    },
    'SGMC': {
        'name': "Singapore Magistrates' Court",
        'input_folders': [os.path.join(FINAL_BASE, "SGMC_2005-2024 Final 1.0")],
        'country': 'SG',
    },
}


# ============================================================================
# LEGAL-AWARE SENTENCE TOKENIZER
# ============================================================================

_legal_tokenizer = None

# Legal abbreviations that should NOT be treated as sentence boundaries.
# NLTK Punkt expects lowercase, without trailing period.
LEGAL_ABBREVIATIONS = {
    # Case citation
    "v",           # versus
    # Paragraph / section references
    "no", "nos", "para", "paras", "s", "ss", "r", "rr",
    "art", "arts", "ch", "cl", "sch", "pt", "reg", "regs",
    "pp", "p", "fn", "ed", "vol", "app",
    # Latin / academic
    "eg", "ie", "cf", "al", "ibid", "op", "cit", "loc", "et", "seq",
    # Company suffixes
    "pte", "sdn", "bhd", "pty", "inc", "ltd", "corp", "plc", "llc", "llp",
    # Titles
    "dr", "mr", "mrs", "ms", "prof", "rev", "hon", "rt", "jr", "sr",
    "gen", "col", "sgt",
    # Legal-specific
    "ex", "re", "dept", "div", "assn", "comm", "dist", "crim", "ors", "anor",
}


def ensure_nltk():
    """Ensure required NLTK data is downloaded."""
    for pkg in ["punkt", "punkt_tab"]:
        try:
            nltk.data.find(f"tokenizers/{pkg}")
        except LookupError:
            nltk.download(pkg, quiet=True)


def get_legal_tokenizer():
    """Get a Punkt sentence tokenizer customised for legal text."""
    global _legal_tokenizer
    if _legal_tokenizer is not None:
        return _legal_tokenizer

    ensure_nltk()
    tokenizer = nltk.data.load("tokenizers/punkt_tab/english.pickle")
    tokenizer._params.abbrev_types.update(LEGAL_ABBREVIATIONS)
    _legal_tokenizer = tokenizer
    return _legal_tokenizer


# Multi-period abbreviations that Punkt can't handle via abbrev_types.
# Temporarily replaced with soft-hyphen placeholders before tokenization.
_MULTI_PERIOD_ABBREVS = [
    (re.compile(r"\be\.g\.", re.IGNORECASE), "e\u00ADg\u00AD"),
    (re.compile(r"\bi\.e\.", re.IGNORECASE), "i\u00ADe\u00AD"),
    (re.compile(r"\bop\.\s*cit\.", re.IGNORECASE), "op\u00ADcit\u00AD"),
    (re.compile(r"\bloc\.\s*cit\.", re.IGNORECASE), "loc\u00ADcit\u00AD"),
    (re.compile(r"\bet\s+al\.", re.IGNORECASE), "et\u00ADal\u00AD"),
    (re.compile(r"\bet\s+seq\.", re.IGNORECASE), "et\u00ADseq\u00AD"),
]

_RESTORE_MAP = [
    ("e\u00ADg\u00AD", "e.g."),
    ("i\u00ADe\u00AD", "i.e."),
    ("op\u00ADcit\u00AD", "op. cit."),
    ("loc\u00ADcit\u00AD", "loc. cit."),
    ("et\u00ADal\u00AD", "et al."),
    ("et\u00ADseq\u00AD", "et seq."),
]


def legal_sent_tokenize(text: str) -> list:
    """Tokenize text into sentences using the legal-aware tokenizer."""
    t = text
    for pattern, replacement in _MULTI_PERIOD_ABBREVS:
        t = pattern.sub(replacement, t)

    tokenizer = get_legal_tokenizer()
    sents = tokenizer.tokenize(t)

    restored = []
    for s in sents:
        for placeholder, original in _RESTORE_MAP:
            s = s.replace(placeholder, original)
        restored.append(s)
    return restored


# ============================================================================
# CITATION STRIPPING FOR READABILITY METRICS
# ============================================================================

LEADING_PARA_NUM = re.compile(r"(?m)^\s*\d{1,3}\.\s+")
BRACKET_CITE = re.compile(r"\[\s*\d{4}\s*\]\s+[A-Z][A-Za-z0-9]+\s+\d+\b")

# Citation pinpoint patterns
PINPOINT_PAREN = re.compile(
    r"\(at\s+(?:"
    r"pages?\s+\d+(?:\s*[-\u2013]\s*\d+)?"
    r"|pp?\.\s*\d+(?:\s*[-\u2013]\s*\d+)?"
    r"|\[\d+\](?:\s*[-\u2013,]\s*\[\d+\])*"
    r")\)"
)
PINPOINT_PARA = re.compile(r"\bat\s+\[\d+\](?:\s*[-\u2013,]\s*\[\d+\])*")
PINPOINT_PAGE = re.compile(
    r"\bat\s+(?:"
    r"pages?\s+\d+(?:\s*[-\u2013]\s*\d+)?"
    r"|pp?\.\s*\d+(?:\s*[-\u2013]\s*\d+)?"
    r")"
)

# Inline footnote/evidence/submission reference patterns
_FOOTNOTE_ABBREVS = (
    r"AEIC|NEs?|AWS|RWS|DCS|PCS|DRS|PRS|SOC|DCC|FNBP|BOA|"
    r"PBOD|DBOD|ROA|ROP|AB|BA|CB|ACB|RCB|DCB|PCB|PA|"
    r"PBD|DBD|JCB|JAEIC"
)
INLINE_FOOTNOTE_REF = re.compile(
    rf"(?:See,?\s+(?:eg|also|generally),?\s+)?"
    rf"(?:{_FOOTNOTE_ABBREVS})"
    rf"(?:\s+\w+)*?"
    rf"\s+at\s+(?:pp?\.?\s*\d[\d\-\u2013\s]*"
    rf"|paras?\s*\d[\d\-\u2013\s]*"
    rf"|\[\d+\](?:\s*[-\u2013,]\s*\[\d+\])*"
    rf"|lines?\s*\d[\d\-\u2013\s]*"
    rf"|pages?\s*\d[\d\-\u2013\s]*)"
    rf"\.?",
    re.IGNORECASE,
)
INLINE_SUBMISSION_REF = re.compile(
    r"(?:Appellant|Respondent|Defendant|Plaintiff|Prosecution|Defence|"
    r"Claimant|Applicant|Petitioner)'?s?\s+"
    r"(?:Written\s+)?(?:Submissions?|Skeletal\s+Arguments?|"
    r"Closing|Reply|Opening)\b[^.]*?\.",
    re.IGNORECASE,
)
INLINE_ROP_REF = re.compile(
    r"Record\s+of\s+[Pp]roceedings\s*\([^)]*\)\s+at\s+[^.]+\.",
    re.IGNORECASE,
)


def remove_para_numbers(text: str) -> str:
    """Remove paragraph numbers (e.g., '1. ', '23. ') from start of lines."""
    return LEADING_PARA_NUM.sub("", text)


def strip_inline_footnote_refs(text: str) -> str:
    """Remove inline footnote/evidence/submission references from text."""
    t = INLINE_FOOTNOTE_REF.sub("", text)
    t = INLINE_SUBMISSION_REF.sub("", t)
    t = INLINE_ROP_REF.sub("", t)
    return t


# Regex-based citation stripping (fast, no spaCy dependency)
# Matches neutral citations: [2017] EWCA Civ 89, [2020] UKSC 1, [2009] HCA 14
_NEUTRAL_CITE = re.compile(
    r'\[\s*\d{4}\s*\]\s+'
    r'(?:UKSC|UKHL|UKPC|EWCA\s+(?:Civ|Crim)|EWHC|UKUT|'
    r'CSOH|CSIH|SGCA|SGHC|SGHCF|SGDC|SGMC|'
    r'HCA|FCAFC|FCA|NSWCA|NSWSC|VSC|VSCA|QCA|QSC|'
    r'NZSC|NZCA|NZHC|'
    r'ABCA|ABQB|ONCA|ONSC|BCCA|BCSC)'
    r'\s+\d+'
)
# Law report citations: [2020] 1 AC 123, [2015] 2 SLR(R) 456
_REPORT_CITE = re.compile(
    r'\[\s*\d{4}\s*\]\s+\d+\s+'
    r'(?:AC|WLR|QB|KB|Ch|Fam|All\s+ER|SLR(?:\(R\))?|MLJ|'
    r'Lloyd|ICR|IRLR|Cr\s+App\s+R|BCLC|BCC|FSR|RPC|STC|WTLR|'
    r'NZLR|DLR|SCR|OR|BCLR|WWR)\s+\d+'
)
# Round-bracket law report: (2005) 224 CLR 123, (2004) 120 LQR 354
_ROUND_CITE = re.compile(
    r'\(\d{4}\)\s+\d+\s+'
    r'(?:CLR|ALR|ALJR|FLR|NSWLR|VR|SASR|WAR|'
    r'US|S\s+Ct|L\s+Ed|F\s+\d[a-z]+|F\s+Supp)\s+\d+'
)


def regex_strip_citations(text: str) -> str:
    """Strip legal citations from text using compiled regexes (fast)."""
    t = _NEUTRAL_CITE.sub("", text)
    t = _REPORT_CITE.sub("", t)
    t = _ROUND_CITE.sub("", t)
    return t


def prepare_text_for_metrics(display_core: str, nlp=None) -> str:
    """
    Full pipeline to prepare text for readability metrics:
    1. Remove paragraph numbers
    2. Remove bracket citations (broad pattern)
    3. Remove neutral/report/round citations (specific patterns)
    4. Remove citation pinpoints
    5. Strip inline footnote/evidence references
    6. Normalize whitespace
    """
    t = remove_para_numbers(display_core)
    t = BRACKET_CITE.sub("", t)
    t = regex_strip_citations(t)
    t = PINPOINT_PAREN.sub("", t)
    t = PINPOINT_PARA.sub("", t)
    t = PINPOINT_PAGE.sub("", t)
    t = strip_inline_footnote_refs(t)
    t = re.sub(r"\s+", " ", t).strip()
    return t


# ============================================================================
# CITATION JURISDICTION CLASSIFICATION
# ============================================================================

# Case citation pattern: [year] REPORTER number  or  (year) volume REPORTER number
CASE_CITATION_PATTERN = re.compile(
    r'\[(\d{4})\]\s*([A-Z][A-Za-z]*(?:\s*\([A-Za-z]+\))?)\s*(\d+)|'
    r'\((\d{4})\)\s*(\d+)\s+([A-Z][A-Za-z]+(?:\s*\d*[A-Za-z]*)?)\s*(\d+)'
)

UK_REPORTERS = {
    'UKHL', 'UKSC', 'UKPC', 'AC', 'WLR', 'QB', 'KB', 'Ch', 'Fam',
    'EWCA', 'EWHC', 'All ER', 'Lloyd', 'ICR', 'IRLR', 'Cr App R',
    'BCLC', 'BCC', 'FSR', 'RPC', 'STC', 'TC', 'WTLR'
}

AU_REPORTERS = {
    'HCA', 'FCAFC', 'FCA', 'NSWCA', 'NSWSC', 'VSC', 'VSCA', 'QCA', 'QSC',
    'WASCA', 'WASC', 'SASC', 'SASFC', 'TASSC', 'TASFC', 'ACTSC', 'ACTCA',
    'CLR', 'ALR', 'ALJR', 'FLR', 'NSWLR', 'VR', 'SASR', 'WAR', 'Qd R',
    'Tas R', 'ACTR', 'NTLR'
}

USA_REPORTERS = {
    'US', 'S Ct', 'L Ed', 'F', 'F 2d', 'F 3d', 'F Supp', 'F Supp 2d',
    'So', 'So 2d', 'NE', 'NE 2d', 'NW', 'NW 2d', 'SE', 'SE 2d',
    'SW', 'SW 2d', 'P', 'P 2d', 'P 3d', 'A', 'A 2d', 'A 3d',
    'Cal', 'NY', 'Ill', 'Tex', 'Mass', 'Pa'
}

CAN_REPORTERS = {
    'SCC', 'SCR', 'FC', 'FCA', 'ONCA', 'ONSC', 'BCCA', 'BCSC',
    'ABCA', 'ABQB', 'DLR', 'OR', 'AR', 'BCLR', 'WWR', 'RFL', 'CCC'
}

IND_REPORTERS = {
    'SCC', 'SCR', 'AIR', 'Bom', 'Cal', 'Mad', 'All', 'Del', 'Kar', 'Ker'
}

NZ_REPORTERS = {
    'NZSC', 'NZCA', 'NZHC', 'NZLR', 'NZAR'
}

SG_REPORTERS = {
    'SGCA', 'SGHC', 'SGHCF', 'SGDC', 'SGMC', 'SLR', 'MLJ', 'SGHCR'
}

EU_REPORTERS = {
    'ECR', 'CMLR', 'EUECJ', 'ECHR'
}


def classify_reporter(reporter: str) -> Optional[str]:
    """Classify a reporter code to its jurisdiction."""
    reporter_clean = reporter.strip().upper()

    # Check each jurisdiction (order: SG first to prioritise local reporters)
    for rep in SG_REPORTERS:
        if rep.upper() in reporter_clean or reporter_clean in rep.upper():
            return 'SG'
    for rep in UK_REPORTERS:
        if rep.upper() in reporter_clean or reporter_clean in rep.upper():
            return 'UK'
    for rep in AU_REPORTERS:
        if rep.upper() in reporter_clean or reporter_clean in rep.upper():
            return 'AU'
    for rep in USA_REPORTERS:
        if rep.upper() in reporter_clean or reporter_clean in rep.upper():
            return 'USA'
    for rep in CAN_REPORTERS:
        if rep.upper() in reporter_clean or reporter_clean in rep.upper():
            return 'CAN'
    for rep in IND_REPORTERS:
        if rep.upper() in reporter_clean or reporter_clean in rep.upper():
            return 'IND'
    for rep in NZ_REPORTERS:
        if rep.upper() in reporter_clean or reporter_clean in rep.upper():
            return 'NZ'
    for rep in EU_REPORTERS:
        if rep.upper() in reporter_clean or reporter_clean in rep.upper():
            return 'EU'

    return 'OTHER'


def count_citations_by_jurisdiction(text: str) -> Dict[str, int]:
    """Count case citations by jurisdiction (total and unique)."""
    counts = {
        'UK': 0, 'AU': 0, 'USA': 0, 'CAN': 0, 'IND': 0,
        'NZ': 0, 'SG': 0, 'EU': 0, 'OTHER': 0, 'total': 0
    }
    unique_counts = {
        'UK': 0, 'AU': 0, 'USA': 0, 'CAN': 0, 'IND': 0,
        'NZ': 0, 'SG': 0, 'EU': 0, 'OTHER': 0, 'total': 0
    }
    seen = set()

    for match in CASE_CITATION_PATTERN.findall(text):
        # Build a normalised citation key for deduplication
        if match[1]:  # bracket form: [year] REPORTER number
            cite_key = f"[{match[0]}] {match[1].strip()} {match[2]}"
            reporter = match[1]
        else:  # round form: (year) volume REPORTER number
            cite_key = f"({match[3]}) {match[4]} {match[5].strip()} {match[6]}"
            reporter = match[5]
        reporter = reporter.strip()
        if len(reporter) >= 2 and not reporter.isdigit():
            jurisdiction = classify_reporter(reporter)
            if jurisdiction:
                counts[jurisdiction] += 1
                counts['total'] += 1
                if cite_key not in seen:
                    seen.add(cite_key)
                    unique_counts[jurisdiction] += 1
                    unique_counts['total'] += 1

    return counts, unique_counts


# ============================================================================
# ACADEMIC REFERENCE DETECTION
# ============================================================================

def count_academic_references(text: str) -> int:
    """Count unique academic references with span-based overlap removal."""

    # Journal article patterns
    journal_patterns = [
        # (year) vol Journal-Name page — require multi-word name with journal keywords
        r'\(\d{4}\)\s+\d+\s+[A-Z][A-Za-z]+(?:\s+[A-Za-z]+)+\s+(?:Law|Legal|Journal|Review|Quarterly|University|Studies)\s+[A-Za-z]*\s*\d*',
        r'\d+\s+(?:Law\s+)?(?:Journal|Review|Quarterly|L\.?\s*J\.?|L\.?\s*Rev\.?|L\.?\s*Q\.?)',
        # UK/general abbreviations (SLR and MLJ removed — they are law reports)
        r'\b(?:LQR|MLR|CLJ|OJLS|CLP|Sing\.?\s*L\.?\s*Rev\.?|SJLS)\b',
        # Full journal names (catch unabbreviated references)
        r'\b(?:Law\s+Quarterly\s+Review|Modern\s+Law\s+Review|Cambridge\s+Law\s+Journal|Oxford\s+Journal\s+of\s+Legal\s+Studies)\b',
        r'\b(?:Yale\s+L\.?\s*J\.?|Harv\.?\s*L\.?\s*Rev\.?|Stan\.?\s*L\.?\s*Rev\.?)\b',
        r'\b(?:Colum\.?\s*L\.?\s*Rev\.?|Mich\.?\s*L\.?\s*Rev\.?|Cornell\s+L\.?\s*Rev\.?)\b',
        # Australian journals (FLR removed — it is Federal Law Reports, a law report series)
        r'\b(?:MULR|UNSWLJ|SydLR|UQLJ|UWALR|AdelLR|MonLR|MelbULawRw)\b',
        r'\b(?:AJLL|ABLR|AIAL\s+Forum|Fed(?:eral)?\s+L(?:aw)?\s+Rev(?:iew)?)\b',
        r'[A-Z][a-z]+,\s*"[^"]+"\s*\(\d{4}\)',
        # J9: SG/Commonwealth journal abbreviations
        r'\b(?:SAcLJ|Mal\.?\s*L\.?\s*R\.?|LMCLQ|JBL|ICLQ|AJCL|Sing\s+L\s+Rev)\b',
        # J10: "Title" (year) Journal — article with quoted title then year then journal
        r'"[^"]{10,}"\s*\(\d{4}\)\s+\d*\s*(?:SAcLJ|LQR|MLR|CLJ|OJLS|Sing\s+L\s+Rev|SJLS|LMCLQ|JBL|ICLQ)',
        # J11: Any quoted title 15+ chars followed by (year)
        r'"[^"]{15,}"\s*\(\d{4}\)',
    ]

    # Book patterns
    book_patterns = [
        r'[A-Z][a-z]+(?:\s+[A-Z][a-z]+)*,\s+[A-Z][A-Za-z\s:]+\([A-Za-z\s]+,\s*\d{4}\)',
        r'[A-Z][a-z]+,\s+[A-Z][A-Za-z\s]+\(\d+(?:st|nd|rd|th)\s+[Ee]d(?:ition)?,\s*\d{4}\)',
        # B3-B4: Named treatises — distinctive single-word author/title names
        r'\b(?:Halsbury|Chitty|Dicey|McGregor|Treitel|Anson|Cheshire|Winfield|Salmond)\b',
        r'\b(?:Snell|Bowstead|Phipson|Archbold|Lewin|Scrutton|Gatley|Keating)\b',
        r'\b(?:MacGillivray|Underhill|Spry|Bennion|Colinvaux|Williston|Corbin)\b',
        r'\b(?:Oppenheim|Brownlie|Pomeroy|Craies|Stroud|Odgers)\b',
        # B7: Multi-word treatise names (author pairs)
        r'\b(?:Clerk\s*&\s*Lindsell|Goff\s*&\s*Jones|Spencer\s+Bower|Smith\s*&\s*Hogan)\b',
        r'\b(?:Mustill\s*&\s*Boyd|Megarry\s*&\s*Wade|Wade\s*&\s*Forsyth|Cross\s*&\s*Tapper)\b',
        r'\b(?:Bullen\s*&\s*Leake|Charlesworth\s*&\s*Percy|de\s+Smith)\b',
        # B8: Common names needing subject context to disambiguate from party names
        r"\bBenjamin'?s?\s+(?:Sale|on\s+Sale)",
        r"\bFleming'?s?\s+(?:Law\s+of\s+Torts|Torts)",
        r"\bGower'?s?\s+(?:Principles|Company|Modern\s+Company)",
        # B9: SG/Commonwealth practitioner works (named titles)
        r'\bSingapore\s+Civil\s+Procedure\b',
        r"\bMallal'?s?\s+Digest\b",
        # B5: Known legal publishers in parentheses with optional edition and year
        r'\((?:Oxford\s+University\s+Press|Cambridge\s+University\s+Press|Hart\s+Publishing|Sweet\s*&\s*Maxwell|LexisNexis|Academy\s+Publishing|Butterworths|Thomson\s+Reuters|Clarendon\s+Press|Stevens|Law\s+Book\s+Co),\s*(?:\d+(?:st|nd|rd|th)\s+[Ee]d(?:ition)?,?\s*)?\d{4}\)',
        # B10: Any university press in parentheses with year (generic catch-all)
        r'\([A-Z][a-z]+(?:\s+[A-Z][a-z]+)?\s+University\s+Press,\s*(?:\d+(?:st|nd|rd|th)\s+[Ee]d(?:ition)?,?\s*)?\d{4}\)',
        # B6: Indian treatises commonly cited in SG courts
        r'\b(?:Ratanlal|Sarkar|Gour)\b(?:\s*&\s*(?:Dhirajlal|Thakore))?\S*\s+(?:Law\s+of|Indian|on\s+)',
    ]

    # Collect all matches with character spans
    matches = []
    for p in journal_patterns:
        for m in re.finditer(p, text, re.IGNORECASE):
            matches.append((m.start(), m.end(), m.group()))
    for p in book_patterns:
        for m in re.finditer(p, text):
            matches.append((m.start(), m.end(), m.group()))

    # Sort by position, longest first for ties — then remove overlaps
    matches.sort(key=lambda x: (x[0], -(x[1] - x[0])))
    used = []
    for s, e, t in matches:
        if any(s < ue and e > us for us, ue, _ in used):
            continue
        used.append((s, e, t))

    return len(used)


# ============================================================================
# TEXT EXTRACTION & METADATA
# ============================================================================

# Neutral citation court codes (for metadata extraction)
NEUTRAL_COURTS = (
    r'(?:SGCA|SGHC|SGHCF|SGDC|SGMC|UKSC|UKHL|UKPC|'
    r'EWCA(?:\s+(?:Civ|Crim))?|EWHC|'
    r'HCA|FCAFC|FCA|NSWCA|NSWSC|VSC|VSCA|QCA|QSC)'
)


def long_path(p: str) -> str:
    """Add Windows long path prefix."""
    if not p.startswith('\\\\?\\'):
        return '\\\\?\\' + os.path.abspath(p)
    return p


def extract_sections(content: str) -> Tuple[str, str, str]:
    """Extract headnotes, core judgment, and footnotes from cleaned txt."""
    headnotes = ""
    core = ""
    footnotes = ""

    hn_match = re.search(
        r'-{10,}\s*\n\s*HEADNOTES\s*\n\s*-{10,}\s*\n(.*?)(?=\n\s*-{10,}\s*\n\s*CORE JUDGMENT)',
        content, re.DOTALL | re.IGNORECASE)
    if hn_match:
        headnotes = hn_match.group(1).strip()

    core_match = re.search(
        r'-{10,}\s*\n\s*CORE JUDGMENT\s*\n\s*-{10,}\s*\n(.*?)(?=\n\s*-{10,}\s*\n\s*FOOTNOTES|\Z)',
        content, re.DOTALL | re.IGNORECASE)
    if core_match:
        core = core_match.group(1).strip()

    fn_match = re.search(
        r'-{10,}\s*\n\s*FOOTNOTES\s*\n\s*-{10,}\s*\n(.*)',
        content, re.DOTALL | re.IGNORECASE)
    if fn_match:
        footnotes = fn_match.group(1).strip()

    return headnotes, core, footnotes


def extract_metadata(content: str, filename: str) -> Dict:
    """Extract metadata from file content and filename."""
    case_match = re.search(r'CASE:\s*(.+)', content)
    full_title = case_match.group(1).strip() if case_match else filename.replace('.txt', '')

    # Extract citation from title
    citation_pat = rf'(\[\d{{4}}\]\s*{NEUTRAL_COURTS}\s*\d+)'
    citation_match = re.search(citation_pat, full_title)
    citation = citation_match.group(1).strip() if citation_match else ""

    # Fallback: try filename
    if not citation:
        citation_match = re.search(citation_pat, filename)
        citation = citation_match.group(1).strip() if citation_match else ""

    # Decision date
    date_match = re.search(r'(?:Decision\s+Date|DATE|Date)\s*:\s*(.+?)(?:\n|$)', content, re.IGNORECASE)
    date = date_match.group(1).strip() if date_match else ""

    if not date:
        headnotes_match = re.search(r'HEADNOTES.*?(?=CORE JUDGMENT)', content, re.DOTALL | re.IGNORECASE)
        if headnotes_match:
            headnotes = headnotes_match.group(0)
            months = r'(?:January|February|March|April|May|June|July|August|September|October|November|December)'
            date_pattern = rf'(\d{{1,2}}\s+{months}\s+\d{{4}})'
            all_dates = re.findall(date_pattern, headnotes, re.IGNORECASE)
            if all_dates:
                date = all_dates[-1]

    # Year from citation or filename
    year_match = re.search(r'\[(\d{4})\]', citation or filename)
    year = year_match.group(1) if year_match else ""

    return {
        'title': full_title,
        'citation': citation,
        'date': date,
        'year': year,
    }


def clean_title(title: str) -> str:
    """Remove citation from title (from [YYYY] onwards)."""
    cleaned = re.sub(rf'\s*\[\d{{4}}\]\s*{NEUTRAL_COURTS}.*$', '', title)
    cleaned = re.sub(r'_cleaned_\d+\.\d+$', '', cleaned)
    return cleaned.strip()


# ============================================================================
# READABILITY METRICS (legal-aware)
# ============================================================================

VOWELS = "aeiouy"


def count_syllables(word: str) -> int:
    """Count syllables in a word using vowel-group counting."""
    w = re.sub(r"[^a-z]", "", word.lower())
    if not w:
        return 0
    syll = 0
    prev_vowel = False
    for ch in w:
        is_vowel = ch in VOWELS
        if is_vowel and not prev_vowel:
            syll += 1
        prev_vowel = is_vowel
    if w.endswith("e") and syll > 1:
        syll -= 1
    return max(1, syll)


def word_count(text: str) -> int:
    """Count words in text."""
    return len(re.findall(r"\b\w+\b", text)) if text.strip() else 0


def sentence_count(text: str) -> int:
    """Count sentences using legal-aware NLTK Punkt tokenizer."""
    if not text.strip():
        return 0
    sents = legal_sent_tokenize(text)
    return max(1, len(sents))


def flesch_kincaid(text: str) -> Tuple[float, float]:
    """Calculate Flesch-Kincaid Grade Level and Reading Ease."""
    if not text.strip():
        return 0.0, 0.0

    words = re.findall(r"\b\w+\b", text)
    wc = max(1, len(words))
    sc = max(1, sentence_count(text))
    syll = sum(count_syllables(w) for w in words)

    fk_grade = 0.39 * (wc / sc) + 11.8 * (syll / wc) - 15.59
    fk_ease = 206.835 - 1.015 * (wc / sc) - 84.6 * (syll / wc)

    return round(fk_grade, 2), round(fk_ease, 2)


def smog_index(text: str) -> float:
    """Calculate SMOG Index (based on polysyllabic word count)."""
    if not text.strip():
        return 0.0

    sents = legal_sent_tokenize(text)
    sc = len(sents)
    if sc == 0:
        return 0.0

    words = re.findall(r"\b\w+\b", text)
    poly = sum(1 for w in words if count_syllables(w) >= 3)

    smog = 1.0430 * math.sqrt(poly * (30 / sc)) + 3.1291
    return round(smog, 2)


def avg_sentence_length(text: str) -> float:
    """Calculate average words per sentence."""
    wc = word_count(text)
    sc = sentence_count(text)
    return round(wc / max(1, sc), 2) if text.strip() else 0.0


# ============================================================================
# MAIN ANALYSIS
# ============================================================================

def analyze_file(filepath: str, court: str, country: str) -> Optional[Dict]:
    """Analyze a single file and return metrics."""
    filename = os.path.basename(filepath)

    try:
        with open(long_path(filepath), 'r', encoding='utf-8') as f:
            content = f.read()
    except Exception as e:
        print(f"  Error reading {filename[:40]}: {e}")
        return None

    content = content.replace('\r\n', '\n').replace('\r', '\n')

    # Extract sections and metadata
    metadata = extract_metadata(content, filename)
    headnotes, core, footnotes = extract_sections(content)
    title = clean_title(metadata['title'])

    # SGCA: remove inline case-header artifacts
    if court == 'SGCA' and metadata.get('citation') and title:
        case_name = re.sub(r'\s+', ' ', title).strip()
        citation = metadata['citation']
        if case_name and citation:
            name_tokens = [re.escape(t) for t in case_name.split() if t]
            if name_tokens:
                name_pattern = r'\s+'.join(name_tokens)
                inline_pat = re.compile(name_pattern + r'.{0,120}?' + re.escape(citation), re.IGNORECASE)
                core = inline_pat.sub('', core)
            core = re.sub(re.escape(citation), '', core)

    # --- Citation & academic counts on RAW core text + footnotes ---
    citation_text = core + "\n" + footnotes if footnotes else core
    citation_counts, unique_counts = count_citations_by_jurisdiction(citation_text)
    academic_refs = count_academic_references(citation_text)

    # --- Readability on PREPARED core text (citations stripped) ---
    core_metrics = prepare_text_for_metrics(core)
    fk_grade, fk_ease = flesch_kincaid(core_metrics)
    smog = smog_index(core_metrics)
    avg_sent = avg_sentence_length(core_metrics)

    return {
        'Title': title,
        'Citation': metadata['citation'],
        'Date': metadata['date'],
        'Year': metadata['year'],
        'Country': country,
        'Court': court,
        'Headnotes_WordCount': word_count(headnotes),
        'Core_WordCount': word_count(core),
        'FK_Grade_Level': fk_grade,
        'FK_Reading_Ease': fk_ease,
        'SMOG': smog,
        'Avg_Sentence_Length': avg_sent,
        'Citations_Total': citation_counts['total'],
        'Citations_Unique': unique_counts['total'],
        'Citations_SG': citation_counts['SG'],
        'Citations_UK': citation_counts['UK'],
        'Citations_AU': citation_counts['AU'],
        'Citations_USA': citation_counts['USA'],
        'Citations_CAN': citation_counts['CAN'],
        'Citations_IND': citation_counts['IND'],
        'Citations_NZ': citation_counts['NZ'],
        'Citations_EU': citation_counts['EU'],
        'Citations_Other': citation_counts['OTHER'],
        'Academic_References': academic_refs,
        'Filename': filename,
    }


def collect_files(input_folders) -> List[str]:
    """Collect .txt files from one or more input folders."""
    if isinstance(input_folders, str):
        input_folders = [input_folders]

    txt_files = []
    for folder in input_folders:
        lp = long_path(folder) if len(folder) > 240 else folder
        if not os.path.exists(lp):
            print(f"  WARNING: Folder not found: {folder}")
            continue
        for root, dirs, files in os.walk(lp):
            for f in files:
                if f.endswith('.txt'):
                    txt_files.append(os.path.join(root, f))
    return sorted(txt_files)


def analyze_folder(input_folders, court: str, country: str,
                   output_xlsx: str, output_csv: str = None) -> None:
    """Analyze all files in folder(s) and export formatted Excel."""
    print("=" * 60)
    print(f"ANALYSIS: {court}")
    print(f"  Readability engine: Legal-aware Punkt tokenizer + spaCy citation stripping")
    print("=" * 60)
    if isinstance(input_folders, str):
        print(f"Input: {input_folders}")
    else:
        for f in input_folders:
            print(f"Input: {f}")
    print(f"Output: {output_xlsx}")
    print()

    print("[INIT] Loading legal sentence tokenizer...")
    get_legal_tokenizer()
    print()

    # Collect files
    txt_files = collect_files(input_folders)
    print(f"[INFO] Found {len(txt_files)} files")
    print()

    if not txt_files:
        print("ERROR: No .txt files found")
        return

    results = []
    errors = []

    for i, filepath in enumerate(txt_files):
        if (i + 1) % 100 == 0:
            print(f"  Progress: {i + 1}/{len(txt_files)}")

        result = analyze_file(filepath, court, country)
        if result:
            results.append(result)
        else:
            errors.append(os.path.basename(filepath))

    print()
    print(f"[EXPORT] Saving Excel with formatting...")

    # Save results
    df = pd.DataFrame(results)
    os.makedirs(os.path.dirname(output_xlsx), exist_ok=True)
    df.to_excel(output_xlsx, index=False, engine='openpyxl')

    # Apply Excel formatting
    wb = load_workbook(output_xlsx)
    ws = wb.active

    dark_grey_fill = PatternFill(start_color="4D4D4D", end_color="4D4D4D", fill_type="solid")
    blue_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    green_fill = PatternFill(start_color="548235", end_color="548235", fill_type="solid")

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    ws.row_dimensions[1].height = 30

    # Column groups: A-F metadata, G-L metrics, M-X citations/academic, Y filename
    metadata_cols = 'ABCDEF'
    metrics_cols = 'GHIJKL'
    citation_cols = 'MNOPQRSTUVWX'

    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border
        if cell.column_letter in metadata_cols or cell.column_letter == 'Y':
            cell.fill = dark_grey_fill
        elif cell.column_letter in metrics_cols:
            cell.fill = blue_fill
        elif cell.column_letter in citation_cols:
            cell.fill = green_fill
        else:
            cell.fill = dark_grey_fill

    column_widths = {
        'A': 40, 'B': 18, 'C': 14, 'D': 8, 'E': 8, 'F': 8,
        'G': 12, 'H': 12, 'I': 12, 'J': 12, 'K': 8, 'L': 12,
        'M': 12, 'N': 12, 'O': 10, 'P': 10, 'Q': 10, 'R': 10,
        'S': 10, 'T': 10, 'U': 10, 'V': 12, 'W': 14, 'X': 14, 'Y': 50,
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.border = thin_border
            if cell.column_letter in 'DEFGHIJKLMNOPQRSTUVWX':
                cell.alignment = Alignment(horizontal="center", vertical="center")

    wb.save(output_xlsx)

    if output_csv:
        df.to_csv(output_csv, index=False)

    print()
    print("=" * 60)
    print("COMPLETE")
    print("=" * 60)
    print(f"  Analyzed: {len(results)}")
    print(f"  Errors:   {len(errors)}")
    print(f"  Excel:    {output_xlsx}")


# ============================================================================
# CLI ENTRY POINT
# ============================================================================

def main():
    parser = argparse.ArgumentParser(description='Analysis Agent for Legal Cases')
    parser.add_argument('--input', '-i', help='Input folder with cleaned .txt files')
    parser.add_argument('--output', '-o', help='Output Excel file path')
    parser.add_argument('--court', '-c',
                        choices=list(COURT_CONFIG.keys()),
                        help='Court to analyze (uses default paths)')
    parser.add_argument('--csv', action='store_true', help='Also output CSV')

    args = parser.parse_args()

    if args.court:
        config = COURT_CONFIG[args.court]
        input_folders = config['input_folders']
        country = config['country']
        court = args.court
        output_xlsx = args.output or os.path.join(OUTPUT_BASE, f"analysis_{court}_Final.xlsx")
        output_csv = output_xlsx.replace('.xlsx', '.csv') if args.csv else None
    elif args.input:
        input_folders = args.input
        court = 'UNKNOWN'
        country = 'UNKNOWN'
        # Auto-detect court from path
        path_upper = args.input.upper()
        for code, cfg in COURT_CONFIG.items():
            if code in path_upper or (code == 'UKHL' and 'HL_' in path_upper):
                court, country = code, cfg['country']
                break
        output_xlsx = args.output or os.path.join(OUTPUT_BASE, f"analysis_{court}_Final.xlsx")
        output_csv = output_xlsx.replace('.xlsx', '.csv') if args.csv else None
    else:
        print("Usage:")
        print("  python -m legal_processor.analyze_cases --court SGCA")
        print("  python -m legal_processor.analyze_cases --court HCA")
        print("  python -m legal_processor.analyze_cases --input [folder] --output [excel]")
        print()
        print("Available courts:", ", ".join(COURT_CONFIG.keys()))
        return

    analyze_folder(input_folders, court, country, output_xlsx, output_csv)


if __name__ == "__main__":
    main()
